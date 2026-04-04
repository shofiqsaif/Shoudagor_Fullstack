import sys
import io

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.utils import get_column_letter

# Load the workbook with formulas (data_only=False)
wb = openpyxl.load_workbook(
    r"C:\Users\Shofiq\Documents\0 work\Shoudagor_Fullstack\Sales March-25.xlsx",
    data_only=False,
)

# Write to a file with explicit UTF-8 encoding
output_file = r"C:\Users\Shofiq\Documents\0 work\Shoudagor_Fullstack\excel_analysis.txt"

with open(output_file, "w", encoding="utf-8") as f:
    f.write("=" * 120 + "\n")
    f.write(f"WORKBOOK SHEETS: {wb.sheetnames}\n")
    f.write("=" * 120 + "\n")

    # Sheets to analyze
    target_sheets = ["SR Proggram", "Daily Report", "Cost & Profit", "Record"]

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            f.write(f"\n[WARNING] Sheet '{sheet_name}' NOT FOUND\n")
            continue

        ws = wb[sheet_name]

        f.write(f"\n{'=' * 120}\n")
        f.write(f"SHEET: '{sheet_name}'\n")
        f.write(f"Dimensions: {ws.dimensions}\n")
        f.write(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}\n")
        f.write(f"Merged Cells: {ws.merged_cells.ranges}\n")
        f.write(f"{'=' * 120}\n")

        # Print ALL non-empty cells with structure
        f.write(f"\n[COMPLETE CELL STRUCTURE]:\n")
        f.write("-" * 120 + "\n")
        for row_num in range(1, (ws.max_row or 0) + 1):
            for col_num in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    col_letter = get_column_letter(col_num)
                    cell_type = (
                        "FORMULA" if str(cell.value).startswith("=") else "VALUE"
                    )
                    f.write(
                        f"  Row {row_num}, {col_letter}: [{cell_type}] {repr(cell.value)}\n"
                    )

        # Print merged cells details
        if ws.merged_cells.ranges:
            f.write(f"\n[MERGED CELLS]:\n")
            f.write("-" * 120 + "\n")
            for merged_range in ws.merged_cells.ranges:
                f.write(f"  {merged_range}\n")

        f.write(f"\n{'=' * 120}\n")
        f.write(f"END OF SHEET: '{sheet_name}'\n")
        f.write(f"{'=' * 120}\n")

    f.write("\nAnalysis complete!\n")

print(f"Analysis written to: {output_file}")
